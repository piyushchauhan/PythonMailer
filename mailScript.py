import smtplib
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email import encoders
import time
import ssl
import imaplib
from os.path import basename
import xlrd
import openpyxl
import csv
import sys


def sendMyMail(fromEmail, toEmails, subject, body):
    msg = MIMEMultipart()
    msg["Subject"] = subject
    msg["From"] = fromEmail
    msg["To"] = toEmails
    # msg["Cc"] = "anwesha.technical@gmail.com"
    msg.attach(MIMEText(body, 'html'))
    s.sendmail(fromEmail, toEmails, msg.as_string())
    print('Mail sent to ', toEmails)


mail_host = '<ENTER HOST SERVER>'
loginID = "<ENTER MAIL ID>"
loginPass = "<ENTER MAIL PASSWORD>"

loc = "/media/pi/Windows/Anwesha/workshopSeriesParticipants.xlsx"
wb = openpyxl.load_workbook(loc)  # , use_iterators=True)
sheet = wb.active

csvfile = open('emailStatus.csv', 'a', newline='\n')
fieldnames = ['email', 'status']
writer = csv.DictWriter(csvfile, delimiter=',', fieldnames=fieldnames)
writer.writeheader()


s = smtplib.SMTP(mail_host, 587)
s.starttls()
s.login(loginID, loginPass)
print('Login success')
fileAttach = 'hack-o-latte-poster.jpeg'

# body = f'''<h4><b>Please circulate among the students.</b></h4><br>
rawbody = f'''Greetings from IIT Patna!<br>
Indian Institute of Technology Patna is going to witness the first edition of Hackathon <b>“Hack O’ Latte”</b>, organised by <a href="https://www.facebook.com/dsciitpatna/">Developer Student Clubs (DSC), IIT Patna</a> in <a href="https://www.facebook.com/anwesha.iitpatna/">Anwesha, annual festival of IIT Patna</a>. Hack O' Latte, is a gathering where programmers code day and night for 36hrs to come up with new and innovative technical solutions. <br><br>

The goal of Hack O'Latte would be to create user friendly products at the end of the event that would benefit people across lots of domains. There will be cutting-edge ideas mostly related to our current industrial needs for the competitors to put their thinking caps on! This Hackthon will also allow participants to come up with their own problem statements. 
We don’t believe in binding the youth and hence there will be no restriction of the technical stack. It will provide a platform <b>for students and startups</b> to come up with new innovative ideas and create new techniques as solutions and participants are free to choose their own problems prevailing in the following domains:<br>
<b><ul>
<li>Agriculture</li>
<li>Socio-economic sector</li>
<li>Industry 4.0</li>
<li>Healthcare</li>
<li>Open innovation</li>
</ul></b><br>
In addition to all these, during the hackathon there will be mentoring sessions by people from technical background  to help the youth to take a more guided and collaborative approach.<br>

<h2><b>PRIZES WORTH ₹1,00,000/-</b><br>
Opportunity to get seed investment of <b>₹10 Lakhs</h2></b><br>

<h2><b>Students are encouraged to register at</b> <a href="https://docs.google.com/forms/u/4/d/e/1FAIpQLSdwUaJ73B9gyZ_t-7wQemnxIqN55quSJ8DyD_u31kDeLEpQbA/viewform/">https://forms.gle/PuYiQgrbXQeeY4Af7</a></h2><br>
<b>Round 1 registeration ends on 2nd Feb 2020</b><br>

<h3>For more details visit <a href="https://anwesha.info/hackathon/">https://anwesha.info/hackathon/</a></h3><br>
In case of any query feel free to contact Piyush Chauhan (+91 9673582517) or Vatsal Singhal (+91 8585992062)<br><br>
Thanks and Regards<br>
Piyush Chauhan<br>
Technical Lead<br>
DSC IIT Patna<br>
+91 9673582517
'''
# recipients = ["1801ee03@iitp.ac.in"]
# recipients.append(str(sheet.cell_value(i, 0)))
# print(str(sheet.cell_value(i, 0)))

# print(recipients)
# for recipient in recipients:
cont = input(
    'Emailing to {} emails. Do you want to continue(y/n)?'.format(sheet.max_row)).lower()
if cont == 'n':
    print('Terminating program...')
    sys.exit()

for i in range(2, sheet.max_row + 2):
    recipient = str(sheet.cell(column=8, row=i).value)
    recipientName = str(sheet.cell(column=3, row=i).value)

    msg = MIMEMultipart()
    msg["Subject"] = "Invitation to participate in Hack O' Latte(Hackathon) at IIT Patna"
    msg["From"] = "dsc@iitp.ac.in"
    msg["To"] = recipient
    # msg["Cc"] = "anwesha.technical@gmail.com"
    body = """Hello {},<br>""".format(recipientName) + rawbody

    msg.attach(MIMEText(body, 'html'))
    with open(fileAttach, "rb") as fil:
        part = MIMEApplication(
            fil.read(),
            Name=basename(fileAttach)
        )
    # After the file is closed
    part['Content-Disposition'] = 'attachment; filename="%s"' % basename(
        fileAttach)
    msg.attach(part)
    text = msg.as_string()
    try:
        s.sendmail(loginID, msg["To"], text)
        print('Mail sent to ', msg["To"])
        sheet.cell(column=13, row=i).value = 'sent;'

        writer.writerow({'email': recipient, 'status': 'sent'})

        imap = imaplib.IMAP4_SSL(mail_host, 993)
        imap.login(loginID, loginPass)
        imap.append('HackOLatte', '\\Seen', imaplib.Time2Internaldate(
            time.time()), text.encode('utf8'))

    except Exception as e:
        print('Error {0} wile sending mail to {1}'.format(e, recipient))
        sheet.cell(column=13, row=i).value = 'not sent;'

        writer.writerow({'email': recipient, 'status': 'not sent'})

wb.save(filename=loc)
csvfile.close()
s.quit()
